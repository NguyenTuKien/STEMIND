from django.shortcuts import render
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.core.files.storage import default_storage
from django.conf import settings
import json
import mimetypes
import os
import PyPDF2
import docx
import openpyxl
from PIL import Image
import io
import base64
from .models import ChatSession, ChatMessage, FileAttachment

def get_message(request):
    message = request.POST.get('message')
    return JsonResponse({'message': message})
# Create your views here.
def chatbot_view(request):
    context = {
        'user': request.user,
    }
    return render(request, 'chatbot.html', context)

@csrf_exempt
@require_http_methods(["POST"])
def chatbot_api(request):
    """Handle chatbot messages and file uploads"""
    try:
        data = json.loads(request.body)
        message = data.get('message', '').strip()
        session_id = data.get('session_id')
        
        if not message:
            return JsonResponse({'success': False, 'error': 'Message is required'})
        
        # Get or create session
        session = get_or_create_session(request, session_id)
        
        # Save user message
        user_message = ChatMessage.objects.create(
            session=session,
            message_type='user',
            content=message
        )
        
        # Generate bot response (simple echo for now)
        bot_response = generate_bot_response(message)
        
        # Save bot message
        bot_message = ChatMessage.objects.create(
            session=session,
            message_type='bot',
            content=bot_response
        )
        
        return JsonResponse({
            'success': True,
            'response': {
                'text': bot_response,
                'session_id': str(session.session_id)
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@csrf_exempt
@require_http_methods(["POST"])
def upload_file(request):
    """Handle file uploads for chatbot"""
    try:
        if 'file' not in request.FILES:
            return JsonResponse({'success': False, 'error': 'No file provided'})
        
        uploaded_file = request.FILES['file']
        session_id = request.POST.get('session_id')
        message_content = request.POST.get('message', '')
        
        # Validate file size (max 10MB)
        max_size = 10 * 1024 * 1024  # 10MB
        if uploaded_file.size > max_size:
            return JsonResponse({'success': False, 'error': 'File too large. Maximum size is 10MB.'})
        
        # Get or create session
        session = get_or_create_session(request, session_id)
        
        # Determine file type
        file_type = get_file_type(uploaded_file.name, uploaded_file.content_type)
        
        # Create user message with file
        user_message = ChatMessage.objects.create(
            session=session,
            message_type='user',
            content=message_content or f"Uploaded file: {uploaded_file.name}"
        )
        
        # Save file attachment
        attachment = FileAttachment.objects.create(
            message=user_message,
            file=uploaded_file,
            original_name=uploaded_file.name,
            file_type=file_type,
            file_size=uploaded_file.size,
            mime_type=uploaded_file.content_type or 'application/octet-stream'
        )
        
        # Extract file content
        file_content = extract_file_content(uploaded_file, file_type)
        
        # Generate bot response based on file content
        bot_response = generate_file_content_response(attachment, file_content, message_content)
        
        # Save bot message
        bot_message = ChatMessage.objects.create(
            session=session,
            message_type='bot',
            content=bot_response
        )
        
        return JsonResponse({
            'success': True,
            'file': {
                'id': attachment.id,
                'name': attachment.original_name,
                'size': attachment.get_file_size_display(),
                'type': attachment.file_type,
                'url': attachment.file.url if attachment.file else None,
                'content_extracted': bool(file_content)
            },
            'response': {
                'text': bot_response,
                'session_id': str(session.session_id)
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def extract_file_content(uploaded_file, file_type):
    """Extract text content from uploaded file"""
    try:
        if file_type == 'document':
            return extract_document_content(uploaded_file)
        elif file_type == 'image':
            return extract_image_content(uploaded_file)
        else:
            return None
    except Exception as e:
        print(f"Error extracting content: {e}")
        return None

def extract_document_content(uploaded_file):
    """Extract text from document files"""
    file_ext = os.path.splitext(uploaded_file.name)[1].lower()
    
    try:
        if file_ext == '.pdf':
            return extract_pdf_content(uploaded_file)
        elif file_ext in ['.doc', '.docx']:
            return extract_word_content(uploaded_file)
        elif file_ext == '.txt':
            return extract_text_content(uploaded_file)
        elif file_ext in ['.xlsx', '.xls']:
            return extract_excel_content(uploaded_file)
        else:
            return None
    except Exception as e:
        print(f"Error extracting document content: {e}")
        return None

def extract_pdf_content(uploaded_file):
    """Extract text from PDF file"""
    try:
        pdf_reader = PyPDF2.PdfReader(uploaded_file)
        text_content = []
        
        for page_num in range(len(pdf_reader.pages)):
            page = pdf_reader.pages[page_num]
            text_content.append(page.extract_text())
        
        return '\n'.join(text_content)
    except Exception as e:
        print(f"Error reading PDF: {e}")
        return None

def extract_word_content(uploaded_file):
    """Extract text from Word document"""
    try:
        doc = docx.Document(uploaded_file)
        text_content = []
        
        for paragraph in doc.paragraphs:
            if paragraph.text.strip():
                text_content.append(paragraph.text)
        
        return '\n'.join(text_content)
    except Exception as e:
        print(f"Error reading Word document: {e}")
        return None

def extract_text_content(uploaded_file):
    """Extract content from text file"""
    try:
        content = uploaded_file.read()
        # Try different encodings
        for encoding in ['utf-8', 'utf-16', 'latin-1', 'cp1252']:
            try:
                return content.decode(encoding)
            except UnicodeDecodeError:
                continue
        return None
    except Exception as e:
        print(f"Error reading text file: {e}")
        return None

def extract_excel_content(uploaded_file):
    """Extract text from Excel file"""
    try:
        workbook = openpyxl.load_workbook(uploaded_file)
        text_content = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_content.append(f"Sheet: {sheet_name}")
            
            for row in sheet.iter_rows(values_only=True):
                row_text = []
                for cell in row:
                    if cell is not None:
                        row_text.append(str(cell))
                if row_text:
                    text_content.append(' | '.join(row_text))
        
        return '\n'.join(text_content)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None

def extract_image_content(uploaded_file):
    """Extract basic info from image file"""
    try:
        image = Image.open(uploaded_file)
        info = {
            'format': image.format,
            'mode': image.mode,
            'size': image.size,
            'width': image.width,
            'height': image.height
        }
        return f"Image info: {info}"
    except Exception as e:
        print(f"Error reading image: {e}")
        return None

def get_or_create_session(request, session_id):
    """Get existing session or create new one"""
    if session_id:
        try:
            session = ChatSession.objects.get(session_id=session_id)
            return session
        except ChatSession.DoesNotExist:
            pass
    
    # Create new session
    session = ChatSession.objects.create(
        user=request.user if request.user.is_authenticated else None,
        title='New Chat'
    )
    return session

def get_file_type(filename, content_type):
    """Determine file type from filename and content type"""
    if content_type:
        if content_type.startswith('image/'):
            return 'image'
        elif content_type.startswith('audio/'):
            return 'audio'
        elif content_type.startswith('video/'):
            return 'video'
        elif content_type in ['application/pdf', 'application/msword', 
                             'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                             'text/plain', 'text/csv',
                             'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             'application/vnd.ms-excel']:
            return 'document'
    
    # Fallback to file extension
    ext = os.path.splitext(filename)[1].lower()
    if ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp']:
        return 'image'
    elif ext in ['.pdf', '.doc', '.docx', '.txt', '.csv', '.xlsx', '.xls', '.pptx']:
        return 'document'
    elif ext in ['.mp3', '.wav', '.ogg', '.m4a']:
        return 'audio'
    elif ext in ['.mp4', '.avi', '.mov', '.wmv']:
        return 'video'
    
    return 'other'

def generate_bot_response(message):
    """Generate bot response using RAG chatbot logic"""
    try:
        # Import the RAG chatbot service
        from File_sharing_platform.services.rag_chatbot_service import get_rag_chatbot_service
        
        # Get RAG chatbot service instance
        rag_service = get_rag_chatbot_service()
        
        if rag_service:
            # Use RAG to answer the question
            response = rag_service.answer_question(message)
            return response
        else:
            # Fallback if RAG service is not available
            return generate_fallback_response(message)
            
    except Exception as e:
        print(f"Lỗi trong RAG chatbot: {e}")
        return generate_fallback_response(message)

def generate_fallback_response(message):
    """Generate fallback response when RAG is not available"""
    message_lower = message.lower()
    
    # STEM Education responses
    if any(word in message_lower for word in ['stem', 'giáo dục', 'education']):
        return '''🔬 **STEM Education** là phương pháp giáo dục tích hợp:

• **S**cience (Khoa học): Vật lý, Hóa học, Sinh học
• **T**echnology (Công nghệ): Máy tính, AI, Robotics  
• **E**ngineering (Kỹ thuật): Thiết kế, xây dựng, sáng tạo
• **M**athematics (Toán học): Logic, phân tích, tính toán

**Lợi ích của STEM:**
• Phát triển tư duy phản biện
• Kỹ năng giải quyết vấn đề
• Sáng tạo và đổi mới
• Chuẩn bị cho thế giới công nghệ'''
    
    # Programming/Coding responses
    elif any(word in message_lower for word in ['code', 'programming', 'lập trình', 'python', 'javascript']):
        return '''💻 **Lập trình** là kỹ năng quan trọng trong thời đại số:

**Ngôn ngữ phổ biến:**
• **Python**: Dễ học, AI/ML, Data Science
• **JavaScript**: Web development, Node.js
• **Java**: Ứng dụng doanh nghiệp, Android
• **C++**: Lập trình hệ thống, game development

**Tại sao nên học lập trình:**
• Tư duy logic và giải quyết vấn đề
• Cơ hội nghề nghiệp rộng mở
• Tự động hóa công việc
• Sáng tạo ứng dụng và website'''
    
    # Science responses
    elif any(word in message_lower for word in ['science', 'khoa học', 'physics', 'chemistry', 'biology']):
        return '''🔬 **Khoa học** là nền tảng hiểu biết về thế giới:

**Các lĩnh vực chính:**
• **Vật lý**: Nghiên cứu vật chất, năng lượng, chuyển động
• **Hóa học**: Cấu trúc, tính chất và biến đổi vật chất
• **Sinh học**: Nghiên cứu sự sống và sinh vật
• **Địa chất**: Nghiên cứu Trái Đất và vũ trụ

**Phương pháp khoa học:**
1. Quan sát hiện tượng
2. Đặt giả thuyết
3. Thiết kế thí nghiệm
4. Thu thập dữ liệu
5. Phân tích và kết luận'''
    
    # Math responses
    elif any(word in message_lower for word in ['math', 'toán', 'toán học', 'mathematics']):
        return '''📐 **Toán học** là ngôn ngữ của khoa học:

**Các lĩnh vực chính:**
• **Đại số**: Phương trình, hàm số, biểu thức
• **Hình học**: Hình dạng, không gian, đo lường
• **Giải tích**: Đạo hàm, tích phân, giới hạn
• **Thống kê**: Phân tích dữ liệu, xác suất

**Ứng dụng thực tế:**
• Kỹ thuật và xây dựng
• Tài chính và kinh tế
• Khoa học máy tính
• Nghiên cứu khoa học'''
    
    # Technology responses
    elif any(word in message_lower for word in ['technology', 'công nghệ', 'ai', 'artificial intelligence', 'machine learning']):
        return '''🤖 **Công nghệ** đang thay đổi thế giới:

**Công nghệ mới nổi:**
• **AI (Trí tuệ nhân tạo)**: Hệ thống thông minh
• **Machine Learning**: Máy học từ dữ liệu
• **Deep Learning**: Mạng neural sâu
• **IoT**: Internet of Things - Kết nối vạn vật
• **Blockchain**: Công nghệ chuỗi khối

**Tác động đến giáo dục:**
• Học tập cá nhân hóa
• Thực tế ảo/tăng cường (VR/AR)
• Nền tảng học trực tuyến
• Phân tích học tập thông minh'''
    
    # Greeting responses
    elif any(word in message_lower for word in ['hello', 'hi', 'xin chào', 'chào']):
        return '''👋 **Xin chào! Tôi là STEMind AI Assistant**

Tôi có thể giúp bạn với:
• 🔬 **STEM Education** - Giáo dục tích hợp
• 💻 **Lập trình** - Python, JavaScript, Java
• 🔬 **Khoa học** - Vật lý, Hóa học, Sinh học  
• 📐 **Toán học** - Đại số, Hình học, Thống kê
• 🤖 **Công nghệ** - AI, Machine Learning, IoT

**Bạn muốn tìm hiểu về chủ đề nào?** 🚀'''
    
    # Help responses
    elif any(word in message_lower for word in ['help', 'giúp', 'hỗ trợ', 'support']):
        return '''🆘 **Tôi có thể giúp bạn với:**

🔬 **Giáo dục STEM**
• Giải thích khái niệm cơ bản
• Phương pháp học tập hiệu quả
• Tài liệu và nguồn học

💻 **Lập trình & Công nghệ**
• Hướng dẫn ngôn ngữ lập trình
• Giải thích thuật toán
• Debug và tối ưu code

📚 **Học tập & Nghiên cứu**
• Giải đáp thắc mắc
• Hướng dẫn làm bài tập
• Tìm kiếm tài liệu

**Hãy đặt câu hỏi cụ thể để tôi có thể hỗ trợ bạn tốt nhất!** 💡'''
    
    # Default response
    else:
        return f'''🤖 **Cảm ơn bạn đã hỏi về: "{message}"**

Tôi là **STEMind AI Assistant** - trợ lý AI chuyên về giáo dục STEM. 

**Tôi có thể giúp bạn với:**
• 🔬 **STEM Education** - Phương pháp giáo dục tích hợp
• 💻 **Lập trình** - Python, JavaScript, Java và nhiều ngôn ngữ khác
• 🔬 **Khoa học** - Vật lý, Hóa học, Sinh học, Địa chất
• 📐 **Toán học** - Đại số, Hình học, Giải tích, Thống kê
• 🤖 **Công nghệ** - AI, Machine Learning, IoT, Blockchain

**Hãy đặt câu hỏi cụ thể hơn để tôi có thể hỗ trợ bạn tốt nhất!** 🚀

*Ví dụ: "Giải thích Machine Learning", "Python là gì?", "Làm thế nào để học STEM?"*'''

def generate_file_content_response(attachment, file_content, user_message):
    """Generate bot response based on file content"""
    file_type = attachment.file_type
    file_name = attachment.original_name
    file_size = attachment.get_file_size_display()
    
    if file_content:
        # Truncate content if too long for display
        content_preview = file_content[:500] + "..." if len(file_content) > 500 else file_content
        
        if file_type == 'document':
            response = f"✅ Tôi đã đọc được nội dung của tài liệu '{file_name}' ({file_size}).\n\n"
            response += f"📄 **Nội dung tài liệu:**\n{content_preview}\n\n"
            
            if user_message:
                response += f"🤔 **Về câu hỏi của bạn:** {user_message}\n\n"
                response += analyze_content_with_question(file_content, user_message)
            else:
                response += "💡 **Tóm tắt:**\n" + summarize_content(file_content)
                response += "\n\n❓ Bạn có câu hỏi gì về nội dung này không?"
            
        elif file_type == 'image':
            response = f"🖼️ Tôi đã nhận được hình ảnh '{file_name}' ({file_size}).\n\n"
            response += f"📊 **Thông tin hình ảnh:**\n{file_content}\n\n"
            
            if user_message:
                response += f"💬 **Về câu hỏi của bạn:** {user_message}\n"
                response += "Tôi có thể thấy thông tin cơ bản về hình ảnh. Bạn có muốn tôi phân tích gì cụ thể không?"
            else:
                response += "❓ Bạn có muốn tôi phân tích điều gì từ hình ảnh này không?"
        
        else:
            response = f"📁 Tôi đã nhận được file '{file_name}' ({file_size}) và trích xuất được một phần nội dung.\n\n"
            response += f"📝 **Nội dung:**\n{content_preview}\n\n"
            response += "❓ Bạn có câu hỏi gì về nội dung này không?"
    
    else:
        # Fallback when content extraction fails
        if file_type == 'image':
            response = f"🖼️ Tôi đã nhận được hình ảnh '{file_name}' ({file_size}). Tôi có thể thấy đây là một file hình ảnh nhưng cần thêm công cụ để phân tích chi tiết nội dung."
        elif file_type == 'document':
            response = f"📄 Tôi đã nhận được tài liệu '{file_name}' ({file_size}). File đã được lưu thành công nhưng tôi gặp khó khăn trong việc đọc nội dung. Có thể do định dạng file hoặc mã hóa đặc biệt."
        else:
            response = f"📁 Tôi đã nhận được file '{file_name}' ({file_size}). File đã được lưu thành công."
        
        if user_message:
            response += f"\n\n💬 **Về câu hỏi của bạn:** {user_message}\nTôi sẽ cố gắng trả lời dựa trên thông tin có sẵn."
    
    return response

def summarize_content(content):
    """Generate a simple summary of the content"""
    if not content:
        return "Không thể tạo tóm tắt."
    
    lines = content.split('\n')
    non_empty_lines = [line.strip() for line in lines if line.strip()]
    
    if len(non_empty_lines) <= 3:
        return content
    
    # Simple summary: first few lines + last line + word count
    summary = []
    summary.extend(non_empty_lines[:2])
    if len(non_empty_lines) > 3:
        summary.append("...")
        summary.append(non_empty_lines[-1])
    
    word_count = len(content.split())
    summary.append(f"\n📊 Tài liệu có {len(non_empty_lines)} dòng, khoảng {word_count} từ.")
    
    return '\n'.join(summary)

def analyze_content_with_question(content, question):
    """Analyze content based on user question"""
    if not content or not question:
        return "Tôi cần thêm thông tin để trả lời câu hỏi này."
    
    question_lower = question.lower()
    content_lower = content.lower()
    
    # Simple keyword matching
    if any(word in question_lower for word in ['tóm tắt', 'summary', 'tổng kết']):
        return summarize_content(content)
    
    elif any(word in question_lower for word in ['tìm', 'find', 'search', 'có']):
        # Extract keywords from question
        keywords = extract_keywords_from_question(question)
        found_lines = []
        
        for line in content.split('\n'):
            if any(keyword.lower() in line.lower() for keyword in keywords):
                found_lines.append(f"• {line.strip()}")
        
        if found_lines:
            return f"Tôi tìm thấy những thông tin liên quan:\n" + '\n'.join(found_lines[:5])
        else:
            return f"Tôi không tìm thấy thông tin cụ thể về '{question}' trong tài liệu này."
    
    else:
        # General response
        return f"Dựa trên nội dung tài liệu, tôi thấy có thông tin liên quan. Bạn có thể đặt câu hỏi cụ thể hơn như 'tìm thông tin về...', 'tóm tắt phần...', v.v."

def extract_keywords_from_question(question):
    """Extract keywords from user question"""
    # Remove common words
    stop_words = ['tìm', 'có', 'là', 'về', 'trong', 'của', 'và', 'với', 'cho', 'từ', 'đến', 'như', 'được', 'sẽ', 'đã', 'find', 'what', 'where', 'when', 'how', 'why', 'is', 'are', 'the', 'a', 'an', 'and', 'or', 'but']
    
    words = question.split()
    keywords = [word.strip('.,!?()[]{}') for word in words if word.lower() not in stop_words and len(word) > 2]
    
    return keywords[:5]  # Return top 5 keywords
